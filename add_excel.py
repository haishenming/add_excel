from openpyxl import Workbook, load_workbook

def open_excel(file_name):
    """ 打开已有 excel文件 """
    wb = load_workbook(file_name)

    return wb

def add_data(filename, datas, start, new_file=False):
    """ 数据写入 excel """
    # 初始化wb
    if new_file:
        wb = Workbook()
    else:
        wb = open_excel(filename)

    base_asc = 64

    ws = wb.active

    row = 0
    for data in datas:
        col = start
        row += 1
        for d in data:
            code = chr(base_asc + col)  # 开始的列数
            t = code + str(row)
            ws[t] = d
            col += 1

    wb.save(filename)

if __name__ == '__main__':
    add_data("tt.xlsx", [[1, 2, 3],[4,5,6]], start=4)